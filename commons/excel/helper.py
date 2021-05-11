# -*- coding: utf-8 -*-

from io import open
from openpyxl import (
    load_workbook,
    Workbook
)

import csv
import os


def get_workbook(filename):
    try:
        name, ext = os.path.splitext(filename)

        if ext == '.csv':
            try:
                wb = Workbook()
                ws = wb.active
                f = open(filename, 'r', encoding='euc-kr')
                reader = csv.reader(f, delimiter=',')

                for row in reader:
                    ws.append(row)
                f.close()
            except Exception:
                wb = Workbook()
                ws = wb.active
                f = open(filename, 'r', encoding='utf-8-sig')
                reader = csv.reader(f, delimiter=',')

                for row in reader:
                    ws.append(row)
                f.close()
        else:
            wb = load_workbook(filename=filename)

        return wb
    except Exception:
        logger.exception('get_workbook')
        return None


def get_header_from_excel(filename, check_header, default='first_line'):
    wb = get_workbook(filename)

    if default in ['first_line']:
        return [cell.value for cell in wb.active[1]]

    for s in wb.active:
        try:
            row = [cell.value for cell in s]
            if set(check_header) < set(row):
                wb.close()
                return list(filter(None.__ne__, row))
        except Exception:
            continue

    wb.close()
    return []
