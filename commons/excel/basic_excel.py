# -*- coding: utf-8 -*-

from .helper import get_workbook

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import os
import traceback


__ALL__ = ['BasicExcelReader', 'BasicExcelWriter']


class BasicExcelReader:
    def __init__(self, filename, no_header=False, skip_lines=0):
        self.filename = filename
        self.rows = []
        self.max_row = 30000
        self.header = None
        self.no_header = no_header
        self.skip_lines = skip_lines
        self.errors = []

        if filename and not os.path.exists(filename):
            logger.debug('file not exists - %s' % filename)

    def set_max_row(self, max_row):
        if max_row > 0:
            self.max_row = max_row

    def read(self):
        self.rows = []

        wb = get_workbook(self.filename)
        max_row = wb.active.max_row

        if self.max_row < max_row:
            self.errors.append('행의수가[%s] 너무 큽니다. %s개의 행만 처리됩니다.' % (max_row, self.max_row))
            max_row = self.max_row

        for i, s in enumerate(wb.active):
            if i > max_row:
                break

            if self.skip_lines > 0:
                self.skip_lines = self.skip_lines - 1
                continue

            try:
                row = [cell.value for cell in s]
                self.rows.append(row)
            except Exception:
                continue

        wb.close()

        if not self.no_header:
            self.header = self.rows.pop(0)
            header = []
            for h in self.header:
                try:
                    if h:
                        header.append(h.strip())
                    else:
                        header.append(h)
                except Exception:
                    header.append(h)
        else:
            header = []
        return header, self.rows


class BasicExcelWriter:
    def __init__(self, filename, rows, header, highlight_field=''):
        self.filename = filename
        self.header = header
        self.rows = rows
        self.highlight_check_idx = -1

        if highlight_field:
            try:
                self.highlight_check_idx = header.index(highlight_field)
            except Exception:
                traceback.print_exc()

    def write(self, title='order'):
        if not self.rows:
            try:
                raise Exception('rows is empty')
            except Exception:
                traceback.print_exc()

            return

        if not self.header:
            try:
                raise Exception('header is empty')
            except Exception:
                traceback.print_exc()

            return

        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = True

        for col in range(0, len(self.header)):
            _ = ws.cell(column=col + 1, row=1, value="{0}".format(self.header[col]))

        for row in range(0, len(self.rows)):
            for col in range(0, len(self.header)):
                try:
                    _ = ws.cell(column=col + 1, row=row + 2, value="{0}".format(self.rows[row][col] if self.rows[row][col] else ''))
                except Exception:
                    logger.exception('BasicExcel write() exception')

            if self.highlight_check_idx > 0 and self.rows[row][self.highlight_check_idx]:
                ws.cell(row=row + 2).fill = redFill

        wb.save(filename=self.filename)

    def append(self, sheet_name=''):
        if not self.rows:
            try:
                raise Exception('rows is empty')
            except Exception:
                traceback.print_exc()

            return

        wb = get_workbook(self.filename)

        try:
            ws = wb.active if sheet_name == '' else wb.get_sheet_by_name(name=sheet_name)
        except Exception:
            logger.exception('')

        max_rows = ws.max_row
        for row in range(0, len(self.rows)):
            for col in range(0, len(self.rows[row])):
                try:
                    _ = ws.cell(column=col + 1, row=row + 1 + max_rows, value="{0}".format(self.rows[row][col] if self.rows[row][col] else ''))
                except Exception:
                    traceback.print_exc()

        wb.save(filename=self.filename)

    def remove_rows(self, sheet_name='', row_numbers=[]):
        wb = get_workbook(self.filename)

        try:
            ws = wb.active if sheet_name == '' else wb.get_sheet_by_name(name=sheet_name)
        except Exception:
            logger.exception('')

        row_numbers.sort()
        while row_numbers:
            number = row_numbers.pop(0)
            ws.delete_rows(number)
            row_numbers = [n - 1 for n in row_numbers]

        wb.save(filename=self.filename)

    def update_cells(self, update_cells, sheet_name=''):
        wb = get_workbook(self.filename)
        try:
            ws = wb.active if sheet_name == '' else wb.get_sheet_by_name(name=sheet_name)
        except Exception:
            traceback.print_exc()
            return

        for c in update_cells:
            try:
                pos = c['position']
                value = c['value']
                ws[pos] = value
            except Exception:
                traceback.print_exc()

        wb.save(filename=self.filename)
